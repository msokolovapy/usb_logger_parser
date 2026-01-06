from itertools import zip_longest
from statistics import mean, median
from datetime import datetime
from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series
from datetime import timedelta
from collections import Counter


FILE_1 = r'N:\GMP Quality Assurance\QA Accessible Documents\Data Logger Downloads\Temperature Mapping\ACPL228\Nov 2025\ACPL149_ACPL228.txt'
FILE_2 = r'N:\GMP Quality Assurance\QA Accessible Documents\Data Logger Downloads\Temperature Mapping\ACPL170\2025\ACPL149_ACPL170.txt'
FILE_3 = r'N:\GMP Quality Assurance\QA Accessible Documents\Data Logger Downloads\Temperature Mapping\ACPL64\2021\ACPL64-ACPL81.txt'
FILE_4 = r'N:\GMP Quality Assurance\QA Accessible Documents\Data Logger Downloads\Temperature Mapping\ACPL156\2020\ACP169_30-03-2020.txt'
FILE_5 = r'N:\GMP Quality Assurance\QA Accessible Documents\Data Logger Downloads\Temperature Mapping\ACPL156\2020\FG 018_30-03-2020.txt'
FILE_6 = r'N:\GMP Quality Assurance\QA Accessible Documents\Data Logger Downloads\Temperature Mapping\ACPL230\Oct 2023\ACPL262.txt'
FILE_7 = r'N:\GMP Quality Assurance\QA Accessible Documents\Data Logger Downloads\Temperature Mapping\ACPL329\Initial 2025\ACPL149_ACPL329.txt'
FILE_8 = r'N:\GMP Quality Assurance\QA Accessible Documents\Data Logger Downloads\Temperature Mapping\ACPL14\2021 Oct\ACPL14-ACPL88.txt'
	
	
class TemperatureData():
	"""This class stores temperature data obtained via Lascar USB data logger.
	Instances of this class can be compared to each other using '==' operator"""
	def __init__(self, temperature_data):
		self._original_data = temperature_data #list of (row_number, date_time_stamp, degrees_celsius) tuples
		self._max = self.select_max_temperature()
		self._min = self.select_min_temperature()
		self._average = self.select_average_temperature()
		self._median = self.select_median_temperature()
		self._data_coll_freq = self.determine_ave_data_coll_frequency()
		self._xlsx_report_data_matrix_size = {'data_width':len(self._original_data[0]),	'data_lenth':len(self._original_data)}

	def __eq__(self, other):
		if not isinstance(other, TemperatureData):
			raise TypeError('Wrong instance type when trying to compare temperature datasets')
		# diff = []
		# #use zip_longest to ensure that data sets of different lengths get zipped properly
		# for i,data in enumerate(zip_longest(self.select_row_temperature(), other.select_row_temperature(), fillvalue = None)):
		# 	data1, data2 = data
		# 	if data1 != data2:
		# 		diff.append((i+1,data1,data2)) #uses i+1 to simplify reading of the below print statement
		# create_print_statement(diff[:100])
		return self._data_coll_freq == other._data_coll_freq
	
	# def select_row_temperature(self):
	# 	return ((x[0],x[2]) for x in self.original_data)
	def select_max_temperature(self):
		return (max(row['celsius'] for row in self.original_data))
	def select_min_temperature(self):
		return (min(row['celsius'] for row in self.original_data))
	def select_average_temperature(self):
		return (round(mean(row['celsius'] for row in self.original_data),2))
	def select_median_temperature(self):
		return (median(row['celsius'] for row in self.original_data))
	
	def determine_ave_data_coll_frequency(self):
		time_diffs = list(self.get_time_diff())  # List of timedelta objects
		avg_timedelta = sum(time_diffs, timedelta()) / len(time_diffs)
		return avg_timedelta

	def get_time_diff(self):
		prev_timestamp = None
		for row in self._original_data:
			if prev_timestamp is not None:
				yield row['date_time'] - prev_timestamp
			prev_timestamp = row['date_time']
	

	@property
	def original_data(self):
		return self._original_data
	@property
	def min(self):
		return self._min
	@property
	def max(self):
		return self._max
	@property
	def average(self):
		return self._average
	@property
	def median(self):
		return self._median
	@property
	def xlsx_report_data_matrix_size(self):
		return self._xlsx_report_data_matrix_size
	@property
	def data_coll_freq(self):
		return self._data_coll_freq


class USBLogger():
	def __init__(self, logger_id, serial_no, data):
		self._data = TemperatureData(data)
		self._id = logger_id
		self._serial_number = serial_no
	
	@classmethod
	def read_from(cls,file_path):
		file_contents = open(file_path)
		header = next(file_contents)
		logger_id, serial_no_idx = obtain_logger_id_from(header)
		serial_no, temperature_data = obtain_serial_numb_temps_from(file_contents, serial_no_idx)
		return cls(logger_id,serial_no,temperature_data)

	def prepare_for_reporting(self):
		#inserts additional data fields to the temp data for better readability when using xlsx report
		copy_temp_data = [tuple(data_point.values()) for data_point in self._data.original_data]
		data_width = self._data.xlsx_report_data_matrix_size['data_width'] 

		header = self.prepare_header(data_width)
		column_names = self.prepare_column_names(data_width)
		
		copy_temp_data.insert(0,column_names)	
		for element in header:
			copy_temp_data.insert(0, element)
		return copy_temp_data

	def prepare_column_names(self, data_width):
		#returns column names for xlsx report such as 'row_numb', 'date_time', 'celsius' based on required data width
		column_names = list(self._data.original_data[0].keys())[:data_width]
		column_names = tuple(column_names)
		return column_names

	def prepare_header(self, data_width):
		header = [self._id, self._serial_number] #logger metadata allows easy logger data identification when multiple logger data traces are overlayed in xlsx report
		padded_header = pad_header_with_Nones(header,data_width)
		return padded_header
	
	@property
	def data(self):
		return self._data
	@property
	def id(self):
		return self._id
	@property
	def serial_number(self):
		return self._serial_number
	

class XLSXReport():
	def __init__(self, *loggers):
		self._loggers = list(loggers)
		self._file_name = self.get_file_name()
		self._wb = Workbook()
		self._x_axis_column = None
		
	def get_file_name(self):
		formatted_today = datetime.now().strftime('%Y-%m-%d')
		return f'{formatted_today}_usb_data_loggers'
	
	def insert_data(self):
		ws = self.wb.active
		ws.title = self.file_name
		for i, usb_logger in enumerate(self.loggers):
			start_col = 10 + (i * 4)  # 3 data columns + 1 skip = 4 spacing
			usb_logger_data = usb_logger.prepare_for_reporting()
			for row_idx, (row_num, date_time, celsius) in enumerate(usb_logger_data, start=1):
				ws.cell(row=row_idx, column=start_col, value=row_num)
				ws.cell(row=row_idx, column=start_col + 1, value=date_time)
				ws.cell(row=row_idx, column=start_col + 2, value=celsius)
		return self.wb
	
	def insert_graph(self):
		ws = self.wb[f'{self.file_name}']
		chart = ScatterChart()
		chart.title = "Temperature Data"
		chart.x_axis.title = "Row"
		chart.y_axis.title = "Temperature (°C)"
		chart.x_axis.delete = False
		chart.y_axis.delete = False
		ws.add_chart(chart, "A1")
		self.wb.save(rf'C:\Users\User\Desktop\Misc\{self._file_name}.xlsx')
	
	def determine_max_row_value(self):
			max_row_value = max(len(logger.data.original_data) for logger in self.loggers)
			return max_row_value
		

	@property
	def loggers(self):
		return self._loggers
	@property
	def file_name(self):
		return self._file_name
	@property
	def wb(self):
		return self._wb

	
#___________________________________________________________________________________________________________________
#	HELPER FUNCTIONS BELOW:

# def obtain_serial_numb_temps_from(file_contents, serial_no_idx):
# 	temperature_data = []
# 	serial_no = None
# 	for line in file_contents:
# 		split_line = line.split(",")

# 		row_numb = int(split_line[0].strip())
# 		date_time = datetime.strptime(split_line[1].strip(),"%Y-%m-%d %H:%M:%S")
# 		celsius = float(split_line[2].strip())
# 		if len(split_line) > 3:
# 			high_alarm = float(split_line[3].strip()) if split_line[3].strip() else None
# 		if len(split_line) > 4:
# 			low_alarm = float(split_line[4].strip()) if split_line[4].strip() else None
# 		if not serial_no:
# 			serial_no = split_line[serial_no_idx].strip() #serial number found once per file in serial_no_idx column
# 		temperature_data.append((row_numb, date_time, celsius,high_alarm, low_alarm))
# 	return serial_no, temperature_data

def obtain_serial_numb_temps_from(file_contents, serial_no_idx):
	temperature_data = []
	serial_no = None
	for line in file_contents:
		split_line = line.split(",")
		if not serial_no:
			serial_no = split_line[serial_no_idx].strip()
		row_data = {
            'row_number': int(split_line[0].strip()),
            'date_time': datetime.strptime(split_line[1].strip(), "%Y-%m-%d %H:%M:%S"),
            'celsius': float(split_line[2].strip())
        }
		if len(split_line) > 3:
			if serial_no == split_line[3].strip():
				pass
			else:
				row_data['high_alarm'] = float(split_line[3].strip()) if split_line[3].strip() else None
			if len(split_line) > 4:
				row_data['low_alarm'] = float(split_line[4].strip()) if split_line[4].strip() else None

		temperature_data.append(row_data)
	return serial_no, temperature_data


def obtain_logger_id_from(header):
	split_header = [field.strip() for field in header.split(',')]
	id = split_header[0].strip()
	serial_no_idx = split_header.index('Serial Number') #position of Serial Number column varies
	return id, serial_no_idx


def create_print_statement(diff_data):
	#formats differences in two data sets to enable easy reading in terminal window
	if not diff_data:
		print('Two temperature data sets are identical')
	else:
		i = len(diff_data)
		j = min(20,i) #max 20 different data points to display in terminal for ease of reading
		trunc_diff_data = diff_data[:j]		
		print(f'There are {i} different data point/s found between two temperature data sets.\nOnly first {j} data point/s are shown: \nPosition|Data set 1|Data set 2')
		for row in trunc_diff_data:
			index, data1, data2 = row
			formatted_data = f'{index:^8}|{str(data1):^10}|{str(data2):^10}'
			print(formatted_data)


@property
def data_coll_freq_identical(loggers):
	if len(set(loggers.data)) == 1:
		return True
	else:
		logger_ids = [logger.id for logger in set(loggers)]
		string_to_display = ', '.join(logger_ids)
		print(f'The following usb loggers have different data collection frequencies: {string_to_display}. \
				Try again by choosing usb loggers with identical data collection frequency to \
				ensure meaningful data comparison')

def pad_header_with_Nones(header,data_width):
	number_Nones = data_width - 1
	padded_header = []
	for logger_metadata in header:
		padded_element = []
		for _ in range(number_Nones):
			padded_element.append(None)
		padded_element.append(logger_metadata)	
		padded_header.append(tuple(padded_element))
	return padded_header
#_______________________________________________________________________________________

# from openpyxl import Workbook
# from openpyxl.chart import (
#     ScatterChart,
#     Reference,
#     Series,
# )

# wb = Workbook()
# ws = wb.active

# rows = [
#     ['Size', 'Batch 1', 'Batch 2'],
#     [2, 40, 30],
#     [3, 40, 25],
#     [4, 50, 30],
#     [5, 30, 25],
#     [6, 25, 35],
#     [7, 20, 40],
# ]

# for row in rows:
#     ws.append(row)

# chart = ScatterChart()
# chart.title = "Scatter Chart"
# chart.style = 13
# chart.x_axis.title = 'Size'
# chart.y_axis.title = 'Percentage'
# chart.x_axis.delete = False
# chart.y_axis.delete = False

# xvalues = Reference(ws, min_col=1, min_row=2, max_row=7)
# for i in range(2, 4):
#     values = Reference(ws, min_col=i, min_row=1, max_row=7)
#     series = Series(values, xvalues, title_from_data=True)
#     chart.series.append(series)

# ws.add_chart(chart, "A10")

# wb.save("C:/Users/User/Desktop/Misc/scatter.xlsx")


# 		



# self.x_axis_column = 18 aka max_row -> class XLSXReport knows max_row_value

# min_row depends on the size of inserted header, which is padded with None to account for empty columns -> Need to take into account data matrix size

# min_column starts at 10 to allow for no overlap with the graph in A1.


	
if __name__ == '__main__':
# 	usb_logger_1 = USBLogger.read_from(FILE_1)
# 	usb_logger_2 = USBLogger.read_from(FILE_2)
# # 	print(f'{usb_logger_1.id} data logger serial number: {usb_logger_1.serial_number}')
# # 	print(f'{usb_logger_2.id} data logger serial number: {usb_logger_2.serial_number}')
# # 	usb_logger_1.data == usb_logger_2.data
# # 	print(f'{usb_logger_1.id} data logger min detected temperature: {usb_logger_1.data.min}')
# # 	print(f'{usb_logger_2.id} data logger max detected temperature: {usb_logger_2.data.max}')
# # 	print(f'{usb_logger_1.id} data logger average detected temperature: {usb_logger_1.data.average}')
# # 	print(f'{usb_logger_2.id} data logger median detected temperature: {usb_logger_2.data.median}')
# 	file = XLSXReport(usb_logger_1,usb_logger_2)
# 	file.insert_data_xlsx()
# 	file.insert_graph_xlsx()

	# file_list = [FILE_1,FILE_2,FILE_3,FILE_4,FILE_5,FILE_6,FILE_7,FILE_8]
	# usb_loggers = [USBLogger.read_from(file) for file in file_list]
	# report = XLSXReport(*usb_loggers)
	# report.insert_data()
	# report.insert_graph()
	
	# frequencies = [logger.data.data_coll_freq for logger in usb_loggers]
	# most_common = Counter(frequencies).most_common()[0][0]

	# for i, logger in enumerate(usb_loggers):
	# 	if logger.data.data_coll_freq != most_common:
	# 		print(f"Logger {logger.id} is an outlier: data collection frequency is {logger.data._data_coll_freq} as compared to most common frequency of {most_common}")


	file_list = [FILE_1,FILE_2,FILE_3,FILE_4,FILE_5,FILE_6,FILE_7,FILE_8]
	usb_loggers = [USBLogger.read_from(file) for file in file_list]
	for logger in usb_loggers:
		prepared_data = logger.prepare_for_reporting()
		print(f'First five rows of Logger {logger.id} data are {prepared_data[:5]}')